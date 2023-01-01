import pandas as pd
import numpy as np
import math
import random
import openpyxl as xl
import os
import sys
import shutil
from datetime import date
import tkinter as tk
from tkinter import messagebox

cur_date = date.today().strftime("%Y-%m-%d")
PCR_name_list = []
sorted_PCR_name_list = []
document_type = ''
senior_row = 0

def get_master():
    return pd.read_excel('1_Master.xlsx', dtype=str)

def get_han_template():
    file_path = 'Templates\Han_Template.xlsx'
    template = xl.load_workbook(file_path)
    han_template = template
    return han_template

def get_delivery_template():
    file_path = 'Templates\Delivery_Template.xlsx'
    template = xl.load_workbook(file_path)
    delivery_template = template
    return delivery_template

def get_phone_template():
    file_path = 'Templates\Phone_Template.xlsx'
    template = xl.load_workbook(file_path)
    phone_template = template
    return phone_template

def change_date (frame, year, month, day):  
    global cur_date
    cur_date = year+'-'+month+'-'+day
    date_label = tk.Label(master=frame, text='=> 바꾼 날짜: '+cur_date, font=('Arial', 15), bg="pale turquoise")
    date_label.pack(side='left')
    return 0

def names_list():
    global senior_row
    master = get_master()
    senior_row = master.loc[master['종익도']=='어르신'].index[0]
    return master['성함'].to_numpy()

def sorted_name_list(PCR_name_list):
    master = get_master()
    full_list = master[['성함','title']].to_numpy()
    title = '0'
    for name in full_list:
        if name[0] in PCR_name_list:
            if name[1] > title:
                sorted_PCR_name_list.append(' ')
                title = name[1]
            sorted_PCR_name_list.append(name[0])
    return sorted_PCR_name_list

def build_label_frame(frame):
    for widget in frame.winfo_children():
        widget.destroy()
    l_row = 0
    l_col = 0
    for name in PCR_name_list:
        label = tk.Label(master=frame, text=name, bg='white', relief='ridge', borderwidth=3, font=('Arial', 13))
        label.grid(row=l_row, column=l_col, padx=5, pady=1)
        l_row += 1
        if l_row > 4:
            l_col += 1
            l_row = 0

def add_remove_PCR_name_list(name, frame):
    if (name in PCR_name_list):
        PCR_name_list.remove(name)
    else:
        PCR_name_list.append(name)
    build_label_frame(frame)
    return 0

def get_name_row(name):
    df = get_master()
    return df[df['성함']==name]

def get_han_data(name):
    data_array = []
    data = get_name_row(name)
    data_array.append(data['성함'].values[0])
    data_array.append(data['생년월일1'].values[0])
    data_array.append(data['성별1'].values[0])
    data_array.append(round(random.uniform(36.1,36.5),1))
    return data_array

def get_delivery_data(name):
    data_array = []
    data = get_name_row(name)
    data_array.append(data['성함'].values[0])
    data_array.append(data['주민등록1'].values[0])
    data_array.append(data['성별2'].values[0])
    return data_array

def get_phone_data(name):
    data_array = []
    data = get_name_row(name)
    data_array.append(data['주민등록2'].values[0])
    data_array.append(data['성함'].values[0])
    data_array.append(data['전화'].values[0])
    return data_array

def open_files(folder, han_excel, han, delivery, phone):
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    elif __file__:
        script_dir = os.path.dirname(__file__)
    os.startfile(os.path.join(script_dir, folder))
    os.startfile(os.path.join(script_dir,delivery))
    os.startfile(os.path.join(script_dir,phone))
    os.startfile(os.path.join(script_dir,han))
    os.startfile(os.path.join(script_dir,han_excel))
    return 0

def create_documents():
    print(cur_date)
    date=cur_date
    print('Documents Processing for '+ date)
    # Create folder with date
    if (os.path.exists('검체/'+date) == False):
        os.mkdir('검체/'+date)
        print('folder exits')
    
    # Make Delivery
    delivery = get_delivery_template()
    delivery_ws = delivery.active
    delivery_date = date.replace('-', '.')
    delivery_save_pathname = '검체/'+date+'/'+'검체의뢰서-'+date+document_type+'.xlsx'
    delivery_ws.cell(row=4, column=7, value=delivery_date)
    row = 0
    for name in PCR_name_list:
        delivery_data = get_delivery_data(name)
        delivery_ws.cell(row=7+row, column=6, value=delivery_data[0])
        delivery_ws.cell(row=7+row, column=7, value=delivery_data[1])
        delivery_ws.cell(row=7+row, column=8, value=delivery_data[2])
        row+=1
    delivery.save(delivery_save_pathname)

    # Make Phone Sheet
    phone = get_phone_template()
    phone_ws = phone.active
    phone_date = date.replace('-', '')
    phone_save_pathname = '검체/'+date+'/'+'고위험-'+date+document_type+'.xlsx'
    row = 0
    for name in PCR_name_list:
        phone_data = get_phone_data(name)
        phone_ws.cell(row=2+row, column=2, value=phone_data[0])
        phone_ws.cell(row=2+row, column=3, value=phone_data[1])
        phone_ws.cell(row=2+row, column=4, value=phone_data[2])
        phone_ws.cell(row=2+row, column=5, value=phone_date)
        row+=1
    phone.save(phone_save_pathname)

    # Make Han Sheet
    han = get_han_template()
    han_ws = han.active
    han_date = date
    if (os.path.exists('검체/'+date+'/han') == False):
        os.mkdir('검체/'+date+'/han')
    han_save_pathname = '검체/'+date+'/han/'+'검사관리대장-'+date+document_type+'.xlsx'
    han_ws.cell(row=1, column=1, value=han_date)
    row = 0
    for name in PCR_name_list:
        han_data = get_han_data(name)
        han_ws.cell(row=1+row, column=2, value=han_data[0])
        han_ws.cell(row=1+row, column=3, value=han_data[1])
        han_ws.cell(row=1+row, column=4, value=han_data[2])
        han_ws.cell(row=1+row, column=5, value=han_data[3])
        row+=1
    han.save(han_save_pathname)

    # Copy hwp template and rename
    shutil.copy('Templates\Han_Template.hwp', '검체/'+date+'/'+'검사관리대장-'+date+document_type+'.hwp')
    
    # Open Files
    open_files( '검체/'+date,han_save_pathname, '검체/'+date+'/'+'검사관리대장-'+date+document_type+'.hwp', delivery_save_pathname, phone_save_pathname)
    print('Documents Created')
    
    return 0

def create_sorted_documents():
    name_list = sorted_name_list(PCR_name_list)
    date=cur_date
    print('Documents Processing for '+ date)
    # Create folder with date
    if (os.path.exists('정리된_검체/'+date) == False):
        os.mkdir('정리된_검체/'+date)
        print('folder exits')
    
    # Make Delivery
    delivery = get_delivery_template()
    delivery_ws = delivery.active
    delivery_date = date.replace('-', '.')
    delivery_save_pathname = '정리된_검체/'+date+'/'+'검체의뢰서-'+date+'.xlsx'
    delivery_ws.cell(row=4, column=7, value=delivery_date)
    row = 0
    for name in name_list:
        if name == ' ':
            delivery_ws.cell(row=7+row, column=6, value=' ')
            delivery_ws.cell(row=7+row, column=7, value=' ')
            delivery_ws.cell(row=7+row, column=8, value=' ')
        else:
            delivery_data = get_delivery_data(name)
            delivery_ws.cell(row=7+row, column=6, value=delivery_data[0])
            delivery_ws.cell(row=7+row, column=7, value=delivery_data[1])
            delivery_ws.cell(row=7+row, column=8, value=delivery_data[2])
        row+=1
    delivery.save(delivery_save_pathname)

    # Make Phone Sheet
    phone = get_phone_template()
    phone_ws = phone.active
    phone_date = date.replace('-', '')
    phone_save_pathname = '정리된_검체/'+date+'/'+'고위험-'+date+'.xlsx'
    row = 0
    for name in name_list:
        if name == ' ':
            phone_ws.cell(row=2+row, column=2, value=' ')
            phone_ws.cell(row=2+row, column=3, value=' ')
            phone_ws.cell(row=2+row, column=4, value=' ')
            phone_ws.cell(row=2+row, column=5, value=' ')
        else:
            phone_data = get_phone_data(name)
            phone_ws.cell(row=2+row, column=2, value=phone_data[0])
            phone_ws.cell(row=2+row, column=3, value=phone_data[1])
            phone_ws.cell(row=2+row, column=4, value=phone_data[2])
            phone_ws.cell(row=2+row, column=5, value=phone_date)
        row+=1
    phone.save(phone_save_pathname)

    # Make Han Sheet
    han = get_han_template()
    han_ws = han.active
    han_date = date
    han_save_pathname = '정리된_검체/'+date+'/검사관리대장-'+date+'.xlsx'
    han_ws.cell(row=1, column=1, value=han_date)
    row = 0
    for name in name_list:
        if name == ' ':
            han_ws.cell(row=1+row, column=2, value=' ')
            han_ws.cell(row=1+row, column=3, value=' ')
            han_ws.cell(row=1+row, column=4, value=' ')
            han_ws.cell(row=1+row, column=5, value=' ')
        else:
            han_data = get_han_data(name)
            han_ws.cell(row=1+row, column=2, value=han_data[0])
            han_ws.cell(row=1+row, column=3, value=han_data[1])
            han_ws.cell(row=1+row, column=4, value=han_data[2])
            han_ws.cell(row=1+row, column=5, value=han_data[3])
        row+=1
    han.save(han_save_pathname)

    # Copy hwp template and rename
    shutil.copy('Templates\Han_Template.hwp', '정리된_검체/'+date+'/'+'검사관리대장-'+date+'.hwp')
    
    return 0

def make_all_documents():
    create_documents()
    create_sorted_documents()
    return 0

def delete_all_names (name_frame, check_frame,management_frame):
    global PCR_name_list
    global sorted_name_list
    global document_type
    document_type = ''
    PCR_name_list = []
    sorted_PCR_name_list = []
    print(PCR_name_list)
    for widget in name_frame.winfo_children():
        widget.destroy()
    for CheckButton in check_frame.winfo_children():
        CheckButton.deselect()
   # build_label_frame(check_frame)

def select_doc_type (option):
    document_types = ['','-어르신', '-종사자']
    global document_type
    if document_type == document_types[option]:
        document_type = document_types[0]
    else: 
        document_type = document_types[option]
    print(document_type)

def add_person(root, name, job, id1,id2, 성별, phone1, phone2, phone3):
   # data for excel
    title_array = ['사무','송용','요양','공익','어르신','시니어', '선택', '기타']
    title = str(title_array.index(job))
    주민등록1 = id1 + '-' + id2
    주민등록2 = id1 + id2
    전화 = phone1 + '-' + phone2 + '-' + phone3
    코로나연락처 = phone1 + phone2 + phone3
    성별1 = '남' if 성별 == 'M' else '여'
    성별2 = '★ M     □ F  ' if 성별 == 'M' else '□ M    ★ F  '
    생년월일1 = id1


    data_array = [title, job, name, 주민등록1, 주민등록2, '', 생년월일1,  성별1, 성별2, 전화, 코로나연락처 ,전화]
   # Pop up for checking information
    confirmation = messagebox.askquestion("확인해주세요", "확인해주세요:\n"+'성함: '+name+'\n'+'주민등록: '+주민등록1+'\n'+'연락처: '+전화+'\n', icon='warning')
    if confirmation == 'yes':
        save_new_person(data_array, root)
    else:
        print ("re-input")

def save_new_person(data_array, root):
    master = get_master()
    master.loc[len(master)] = data_array
    print(master.head())
    master = master.sort_values(['title', '성함'])
    master.reset_index(drop=True, inplace=True)
    print(master.head())
    
    # Save to Excel
    book = xl.load_workbook('1_Master.xlsx')
    writer = pd.ExcelWriter('1_Master.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    master.to_excel(writer, "Master", index=False)
    writer.save()

    # Reload name list
    root.destroy()
    runTk()
    return 0

def runTk():
    root = tk.Tk()
    root.title('검체 만들기')
    appName = tk.Label(text="검체 만들기 프로그램", bg='deep sky blue')
    root.configure(bg='deep sky blue')
    root.iconbitmap('ace.ico')
    appName.pack()
    
    # Add 4 Frames
    date_frame = tk.Frame(master=root, height=100, bg="pale turquoise")
    date_frame.pack(fill=tk.X)
    number_frame = tk.Frame(master=root, height=100, bg="light cyan")
    number_frame.pack(fill=tk.X, padx=258)
    list_frame = tk.Frame(master=root, height=100, bg="white")
    list_frame.pack(fill=tk.X, padx=258)
    all_names_frame = tk.Frame(master=root, height=100, bg="snow")
    all_names_frame.pack(fill=tk.X)
    add_frame = tk.Frame(master=root, height=25, bg="light cyan")
    add_frame.pack(fill=tk.X)
    management_frame = tk.Frame(master=root, height=25, bg="sky blue")
    management_frame.pack(fill=tk.X)
    submission_frame = tk.Frame(master=root, height=25, bg="deep sky blue")
    submission_frame.pack(fill=tk.X)

    # Add Date
    date_label = tk.Label(master=date_frame, text='오늘 날짜: '+cur_date, font=('Arial', 15), bg="pale turquoise")
    date_label.pack(side='left')

    # Modify Date
    modify_date = tk.Button(master=date_frame, text='날짜 바꾸기', command=lambda:change_date(date_frame ,year_var.get(), month_var.get(), day_var.get()), font=('Arial', 10), bg="pale turquoise")
    modify_date.pack(side='right')

    year=['2022', '2023', '2024']
    year_var = tk.StringVar(root)
    year_var.set('2022')
    year = tk.OptionMenu(date_frame, year_var, *year)
    year_label = tk.Label(master=date_frame, text='년', font=('Arial', 10), bg="pale turquoise")
    year_label.pack(side='right')
    year.pack(side='right')
    month=['01','02','03','04','05','06','07','08','09','10','11','12']
    month_var = tk.StringVar(root)
    month_var.set(' ')
    month = tk.OptionMenu(date_frame, month_var, *month)
    month_label = tk.Label(master=date_frame, text='월', font=('Arial', 10), bg="pale turquoise")
    month_label.pack(side='right')
    month.pack(side='right')
    day=['01','02','03','04','05','06','07','08','09','10', '11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30', '31']
    day_var = tk.StringVar(root)
    day_var.set(' ')
    day = tk.OptionMenu(date_frame, day_var, *day)
    day_label = tk.Label(master=date_frame, text='일', font=('Arial', 10), bg="pale turquoise")
    day_label.pack(side='right')
    day.pack(side='right')

    # Add Default Labels
    for i in range (8):
        label = tk.Label(master=number_frame, text=str(i+1), font=('Arial', 15), justify='center', bg="light cyan")
        label.grid(row=0, column=i, padx=26)

    # Add name buttons to all_names_frame
    name_list = names_list()
    print(name_list)
    row = math.ceil(len(name_list)/12)
    b_row = 0
    b_col = 0
    i = 0
    label = tk.Label(master=all_names_frame, text='종사자',bg="snow", font=('Arial', 14),)
    label.grid(row=b_row, column=b_col)
    b_row += 1
    for name in names_list():
        button = tk.Checkbutton(master=all_names_frame, text=name, bg="snow", font=('Arial', 12), command=lambda name=name: add_remove_PCR_name_list(name, list_frame))
        button.grid(row=b_row, column=b_col)
        b_col += 1
        i += 1
        if (b_col > 12):
            b_row += 1
            b_col = 0
        if (senior_row == i):
            b_col = 0
            b_row += 1
            label = tk.Label(master=all_names_frame, text='어르신',bg="snow", font=('Arial', 14),)
            label.grid(row=b_row, column=b_col)
            b_row += 1

    # Management Options
    delete_all_btn = tk.Button(master=management_frame, fg='white', bg='red', text='모두 삭제하기', font=('Arial', 15), command=lambda:delete_all_names(list_frame, all_names_frame, management_frame))        
    delete_all_btn.pack(side='right')
    어르신_btn = tk.Checkbutton(master=management_frame, bg='sky blue', text='어르신', font=('Arial', 15), command=lambda:select_doc_type(1))        
    종사자_btn = tk.Checkbutton(master=management_frame, bg='sky blue', text='종사자', font=('Arial', 15), command=lambda:select_doc_type(2))        
    종사자_btn.pack(side='left')
    어르신_btn.pack(side='left')

    # Add New Person
    job_array = ['사무','송용','요양','공익','어르신','시니어','기타']
    title = 5
    name = ''
    job = tk.StringVar(root) # OptionMenu
    job.set('선택')
    주민등록1 = tk.StringVar(root)
    주민등록2 = tk.StringVar(root)
    성별 = tk.StringVar(root) # OptionMenu
    성별.set('선택')
    성별_array = ['M','F']
    phone1 = tk.StringVar(root)
    phone2 = tk.StringVar(root)
    phone3 = tk.StringVar(root)
    new_label = tk.Label(master=add_frame, text='사람 추가하기',font=('Arial', 15),bg="light cyan")
    new_label.pack(side='top')
    name_label = tk.Label(master=add_frame, text='성함',font=('Arial', 12),bg="light cyan")
    name_textbox = tk.Text(master=add_frame, height=1, width=7)
    name_label.pack(side='left')
    name_textbox.pack(side='left')
    주민등록_label = tk.Label(master=add_frame, text='주민등록',font=('Arial', 12),bg="light cyan")
    주민등록_label.pack(side='left')
    주민등록1_textbox = tk.Text(master=add_frame, height=1, width=7)
    주민등록1_textbox.pack(side='left')
    dash_label = tk.Label(master=add_frame, text='-',font=('Arial', 12),bg="light cyan")
    dash_label.pack(side='left')
    주민등록2_textbox = tk.Text(master=add_frame, height=1, width=7)
    주민등록2_textbox.pack(side='left')
    phone_label = tk.Label(master=add_frame, text='연락처',font=('Arial', 12),bg="light cyan")
    phone_label.pack(side='left')
    phone_textbox1 = tk.Text(master=add_frame, height=1, width=3)
    phone_textbox1.pack(side='left')
    dash_label = tk.Label(master=add_frame, text='-',font=('Arial', 12),bg="light cyan")
    dash_label.pack(side='left')
    phone_textbox2 = tk.Text(master=add_frame, height=1, width=4)
    phone_textbox2.pack(side='left')
    dash_label = tk.Label(master=add_frame, text='-',font=('Arial', 12),bg="light cyan")
    dash_label.pack(side='left')
    phone_textbox3 = tk.Text(master=add_frame, height=1, width=4)
    phone_textbox3.pack(side='left')
    job_label = tk.Label(master=add_frame, text='종익도:',font=('Arial', 12),bg="light cyan")
    성별_label = tk.Label(master=add_frame, text='성별:',font=('Arial', 12),bg="light cyan")
    job_dropdown = tk.OptionMenu(add_frame, job, *job_array)
    성별_dropdown = tk.OptionMenu(add_frame, 성별, *성별_array)
    job_label.pack(side='left')
    job_dropdown.pack(side='left')
    성별_label.pack(side='left')
    성별_dropdown.pack(side='left')
    #add_button = tk.Button(master=add_frame, text='추가하기', command=lambda:add_person(name_textbox.get('1.0','end-1c')))
    add_button = tk.Button(master=add_frame, text='추가하기', font=('Arial', 13), bg="pale turquoise", command=lambda:add_person(root, name_textbox.get('1.0','end-1c'), job.get(), 주민등록1_textbox.get('1.0','end-1c'),주민등록2_textbox.get('1.0','end-1c'), 성별.get(), phone_textbox1.get('1.0','end-1c'), phone_textbox2.get('1.0','end-1c'), phone_textbox3.get('1.0','end-1c')))
    add_button.pack(side='right')

    # Submit Button
    submission = tk.Button(master=submission_frame, bg='deep sky blue', text='파일 만들기', font=('Arial', 15), command=lambda:make_all_documents())
    submission.pack()

    root.mainloop()
    return 0

def main():
    runTk()
    return 0 

main()